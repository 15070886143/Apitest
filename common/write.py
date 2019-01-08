# -*- coding:utf-8 -*-

"""
@version: 1.0
@author: fky
@site:
@software: PyCharm
@file: sendRequests.py
@time: 2018/3/24 11:40
"""
from openpyxl import load_workbook
import openpyxl
from Apitest.common import updata_sheetnames
#设置两个参数，文件路径
def copy_excel(excelpath1, excelpath2):
    '''复制excek，把excelpath1数据复制到excelpath2'''
    wb2 = openpyxl.Workbook()
    wb2.save(excelpath2)
    # 读取数据
    wb1 = openpyxl.load_workbook(excelpath1)
    wb2 = openpyxl.load_workbook(excelpath2)
    sheets1 = wb1.sheetnames
    print('sheets1',sheets1)
    sheets2 = wb2.sheetnames
    print('sheets2', sheets2)
    sheet1 = wb1[sheets1[0]]
    print('sheet1',sheet1)
    sheet2 = wb2[sheets2[0]]
    print('sheet2', sheet2)
    max_row = sheet1.max_row         # 最大行数
    print('max_row',max_row)
    max_column = sheet1.max_column   # 最大列数
    print('max_column',max_column)
    for m in list(range(1,max_row+1)):
        for n in list(range(97,97+max_column)):   # chr(97)='a'
            n = chr(n)                            # ASCII字符
            i ='%s%d'% (n, m)                     # 单元格编号
            cell1 = sheet1[i].value               # 获取data单元格数据
            sheet2[i].value = cell1               # 赋值到test单元格
    wb2.save(excelpath2)                 # 保存数据
    wb1.close()                          # 关闭excel
    wb2.close()
class Write_excel(object):
    '''修改excel数据'''
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active  # 激活sheet

    def write(self, row_n, col_n, value):
        '''写入数据，如(2,3，"hello"),第二行第三列写入数据"hello"'''
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)



if __name__ == "__main__":
    copy_excel('../data/textcase01.xlsx','../case/demo-text.xlsx')
    wt = Write_excel('../case/demo-text.xlsx')
    updata_sheetnames.updata_sheetnames()
    print('ok')